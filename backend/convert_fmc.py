"""
Convert the FMC OTI List export (.xlsx) into the FreightLink import CSV format.

Usage:
    python3 convert_fmc.py "OTI List ....xlsx" carriers_import.csv [--include-ff] [--us-only]

Stores a fuller record than the UI shows. Maps FMC columns -> import columns:
    company_name    <- Name
    contact_name    <- QI 1 (first Qualifying Individual)
    qi_title        <- QI 1 Title
    email           <- (blank; FMC does not publish emails)
    phone           <- Phone
    fax             <- Fax
    country         <- Nation
    city / state    <- City / State
    zip / street    <- Zip / Street 1 (+ Street 2)
    lanes           <- (blank; FMC does not track trade lanes)
    fmc_id          <- Organization Number
    license_number  <- License Number
    trade_name      <- NVOCC DBA Name 1 (fallback FF DBA Name 1)
    carrier_type    <- NVOCC / FF / NVOCC + FF (from the NVOCC & FF flags)
    renewal_date    <- Renewal Date (YYYY-MM-DD)

By default only NVOCCs are included. --include-ff also includes forwarders;
--us-only keeps only United States records.
"""

import csv
import sys
import datetime
import openpyxl

OUT_FIELDS = ['company_name', 'contact_name', 'email', 'phone', 'country', 'lanes',
              'fmc_id', 'trade_name', 'license_number', 'city', 'state', 'zip',
              'street', 'carrier_type', 'qi_title', 'renewal_date', 'fax']


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    include_ff = '--include-ff' in sys.argv
    us_only = '--us-only' in sys.argv
    if len(args) < 2:
        print('Usage: python3 convert_fmc.py input.xlsx output.csv [--include-ff] [--us-only]')
        sys.exit(1)

    inp, outp = args[0], args[1]
    wb = openpyxl.load_workbook(inp, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    next(rows)  # banner row
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}

    def raw(row, name):
        i = idx.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    def g(row, name):
        v = raw(row, name)
        return ('' if v is None else str(v)).strip()

    def gdate(row, name):
        v = raw(row, name)
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.strftime('%Y-%m-%d')
        return ('' if v is None else str(v)).strip()[:10]

    total = kept = 0
    seen = set()
    out_rows = []
    for row in rows:
        if not any(row):
            continue
        total += 1
        is_nvocc = g(row, 'NVOCC').upper() == 'NVOCC'
        is_ff = g(row, 'FF').upper() == 'FF'
        if not is_nvocc and not (include_ff and is_ff):
            continue
        if us_only and g(row, 'Nation').upper() != 'UNITED STATES':
            continue

        name = g(row, 'Name')
        if not name:
            continue
        org = g(row, 'Organization Number')
        key = org or name.lower()
        if key in seen:
            continue
        seen.add(key)

        if is_nvocc and is_ff:
            ctype = 'NVOCC + FF'
        elif is_nvocc:
            ctype = 'NVOCC'
        else:
            ctype = 'FF'

        street = ' '.join(p for p in [g(row, 'Street 1'), g(row, 'Street 2')] if p)
        trade = g(row, 'NVOCC DBA Name 1') or g(row, 'FF DBA Name 1')

        out_rows.append({
            'company_name': name,
            'contact_name': g(row, 'QI 1'),
            'email': '',
            'phone': g(row, 'Phone'),
            'country': g(row, 'Nation'),
            'lanes': '',
            'fmc_id': org,
            'trade_name': trade,
            'license_number': g(row, 'License Number'),
            'city': g(row, 'City'),
            'state': g(row, 'State'),
            'zip': g(row, 'Zip'),
            'street': street,
            'carrier_type': ctype,
            'qi_title': g(row, 'QI 1 Title'),
            'renewal_date': gdate(row, 'Renewal Date'),
            'fax': g(row, 'Fax'),
        })
        kept += 1

    with open(outp, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=OUT_FIELDS)
        w.writeheader()
        w.writerows(out_rows)

    print(f'Scanned rows:     {total}')
    print(f'Written carriers: {kept}')
    print(f'Output: {outp}')


if __name__ == '__main__':
    main()
